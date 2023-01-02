# -*- coding: utf-8 -*-


from rich import print
from windrose import WindroseAxes
import openpyxl
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
import pandas as pd
import datetime
from datetime import datetime
from collections import Counter
from tabulate import tabulate
import matplotlib.cm as cm
from math import pi
from matplotlib.pyplot import figure, show, rc
from IPython.display import Image
#необхідні модулі


def Region():

    print("[bold blue]Виберіть регіон із списку:[/bold blue]")
    print("[italic yellow]1 - Дніпропетровськ[/italic yellow]")
    print("[italic yellow]2 - Донецьк[/italic yellow]")
    print("[italic yellow]3 - Івано Франківськ[/italic yellow]")
    print("[italic yellow]4 - Харків[/italic yellow]")
    print("[italic yellow]5 - Кривий Ріг[/italic yellow]")
    print("[italic yellow]6 - Київ[/italic yellow]")
    print("[italic yellow]7 - Луганськ[/italic yellow]")
    print("[italic yellow]8 - Львів[/italic yellow]")
    print("[italic yellow]9 - Одеса[/italic yellow]")
    print("[italic yellow]10 - Сімферополь[/italic yellow]")

    get_region = int(input("Введіть індекс регіону: "))

    if get_region == 1:
        print("[italic green]вибрано Дніпропетровськ[/italic green]")
        path = r"Database/Dnipropetrovsk/"
    elif get_region == 2:
        print("[italic green]вибрано Донецьк[/italic green]")
        path = r"Database/Donetsk/"
    elif get_region == 3:
        print("[italic green]вибрано Івано Франківськ[/italic green]")
        path = r"Database/Ivano-Frankivsk/"
    elif get_region == 4:
        print("[italic green]вибрано Харків[/italic green]")
        path = r"Database/Kharkiv/"
    elif get_region == 5:
        print("[italic green]вибрано Кривий Ріг[/italic green]")
        path = r"Database/Kryvyi_Rih/"
    elif get_region == 6:
        print("[italic green]вибрано Київ[/italic green]")
        path = r"Database/Kyiv/"
    elif get_region == 7:
        print("[italic green]вибрано Луганськ[/italic green]")
        path = r"Database/Luhansk/"
    elif get_region == 8:
        print("[italic green]вибрано Львів[/italic green]")
        path = r"Database/Lviv/"
    elif get_region == 9:
        print("[italic green]вибрано Одеса[/italic green]")
        path = r"Database/Odesa/"
    elif get_region == 10:
        print("[italic green]вибрано Сімферополь[/italic green]")
        path = r"Database/Simferopol/"
    else:
        print("[bold red][!] Помилка при виборі регіону, спробуйте знову[/bold red]")

    return path


def Calendar(path):
    
    print("[bold blue]Виберіть номер місяця за яким надати інформацію: [/bold blue]")
    global month_number
    month_number = input("Введіть календарний номер місяця: ")

    dataset = f"{path}2012-{month_number}.xlsx" # створюємо шлях до файла із яким будемо працювати

    return dataset

def Temperature(dataset):

    book = openpyxl.open(dataset, read_only = True) # Відкриваємо Excel файл тільки для читання
    sheet = book.active # вибираємо сторінку

    data = []
    temperature = []
    month = month_number

    if len(month) == 1:
        month = "0" + month
   
    for row in range(2, sheet.max_row):
        
        day = str(sheet[row][0].value)

        if len(day) == 1:
            day = "0" + day

        hour = str(sheet[row][1].value.hour)

        if len(hour) == 1:
            hour = "0" + hour

        minute = str(sheet[row][1].value.minute)

        if len(minute) == 1:
            minute= "0" + minute

        numbers = "2012" + month + day + hour + minute + "00" # збираємо усі дані про дату та зєднюємо їх у 1 рядок
        date = datetime.strptime(numbers, "%Y%m%d%H%M%S") # Переводимо рядок з датою у тип datetime
        
        data.append(date) # записуємо значення дат в список
        temperature.append(sheet[row][2].value) # записуємо значення температур в список

    #print(data, temperature)
    f = plt.figure() # Створюємо нову фігуру
    f.set_size_inches(30, 40) # Задаємо розміри
    plt.plot(data,temperature) # Додаємо дані для лінійного графіка
    plt.gcf().autofmt_xdate() # Перетворюємо дату у підходящий тип даних
    plt.title("Графік температури") # Назва графіка
    plt.xlabel("Дата") # Назва осі Х
    plt.ylabel("Температура (°C)") # назва осі У
    plt.show() # демонструємо графік
    f.savefig(r"report/task_2.2_plot.pdf") # Зберігаємо графік у пдф за шляхом


def temperature_regime(dataset): 

    book = openpyxl.open(dataset, read_only = True)
    sheet = book.active

    temperature = []

    for row in range(2, sheet.max_row):

        temperature.append(sheet[row][2].value)

    elements = Counter(temperature) # Получаємо словник із значеннями та їхньою кількістю

    values = []
    numbers = []

    for key in elements:

        values.append(key) # Із словника створюємо список із значеннями
        numbers.append(elements[key] / 2) # Із словника створюємо список із кількістю цих значень

    #print(values, numbers)
    values1 = ["Температура"] + values # створюємо значення для таблиці із "заголовком"
    numbers1 = ["Значення"] + numbers 

    data = [
        values1,
        numbers1
    ] # значення таблиці

    fig, ax = plt.subplots()
    print("[bold blue]Теблиця залежності температур[/bold blue]")
    print(tabulate(data, tablefmt="grid")) # створюємо та виводимо таблицю із стилем

    ax.bar(values, numbers) # створюємо стовпцеву діаграму
    plt.title("діаграмa тривалості температурних режимів")
    plt.xlabel("Температура (°C)")
    plt.ylabel("Час (год)")
    ax.set_facecolor("seashell")
    fig.set_facecolor("floralwhite")
    fig.set_figwidth(len(values)) # Задаємо розміри діаграми за кількістю даних
    fig.set_figheight(len(numbers))
    fig.savefig(r"report/task_2.3_plot.pdf")

    plt.show()


def Rose_winds(dataset):

    book = openpyxl.open(dataset, read_only = True)
    sheet = book.active

    radii = []
    speed = []

    for row in range(2, sheet.max_row):

        if sheet[row][3].value == "Северный": # переводимо напрямки у градуси
            radii.append(0)
            speed.append(sheet[row][4].value)

        elif sheet[row][3].value == "С-В":
            radii.append(45)
            speed.append(sheet[row][4].value)

        elif sheet[row][3].value == "Восточный":
            radii.append(90)
            speed.append(sheet[row][4].value)

        elif sheet[row][3].value == "Ю-В":
            radii.append(135)
            speed.append(sheet[row][4].value)

        elif sheet[row][3].value == "Южный":
            radii.append(180)
            speed.append(sheet[row][4].value)

        elif sheet[row][3].value == "С-З":
            radii.append(225)
            speed.append(sheet[row][4].value)

        elif sheet[row][3].value == "Западный":
            radii.append(270)
            speed.append(sheet[row][4].value)

        elif sheet[row][3].value == "Ю-З":
            radii.append(315)
            speed.append(sheet[row][4].value)

        elif sheet[row][3].value == "Переменный": # для змінного записуємо усі напрямки
            radii.append(0)
            speed.append(sheet[row][4].value)
            radii.append(45)
            speed.append(sheet[row][4].value)
            radii.append(90)
            speed.append(sheet[row][4].value)
            radii.append(135)
            speed.append(sheet[row][4].value)
            radii.append(180)
            speed.append(sheet[row][4].value)
            radii.append(225)
            speed.append(sheet[row][4].value)
            radii.append(270)
            speed.append(sheet[row][4].value)
            radii.append(315)
            speed.append(sheet[row][4].value)

    #print(speed, radii)

    ax = WindroseAxes.from_ax()
    ax.bar(radii, speed, normed=True, opening=0.8, edgecolor='white') # створюємо векторну діаграму
    ax.set_legend()

    plt.savefig(r"report/task_2.4_plot.pdf") # зберігаємо

    show()


def wind_activity(dataset):

    book = openpyxl.open(dataset, read_only = True)
    sheet = book.active

    speed = []

    for row in range(2, sheet.max_row):

        speed.append(sheet[row][4].value) # Получаємо швидкість вітрів

    elements = Counter(speed) # Получаємо словник із значеннями та їхньою кількістю

    values = []
    numbers = []

    for key in elements:

        values.append(key) # Із словника створюємо список із значеннями
        numbers.append(elements[key] / 2) # Із словника створюємо список із кількістю цих значень

    values1 = ["Швидкість вітру (м/с)"] + values # створюємо значення для таблиці із "заголовком"
    numbers1 = ["Тривалість (год)"] + numbers 

    data = [
        values1,
        numbers1
    ] # значення таблиці

    fig, ax = plt.subplots()
    print("[bold blue]Теблиця залежності вітрів[/bold blue]")
    print(tabulate(data, tablefmt="grid")) # створюємо та виводимо таблицю із стилем

    ax.bar(values, numbers) # створюємо стовпцеву діаграму
    plt.title("діаграмa вітрового потенціалу за швидкостями год")
    plt.xlabel("Швидкість вітру (м/с)")
    plt.ylabel("Tривалість (год)")
    ax.set_facecolor("seashell")
    fig.set_facecolor("floralwhite")
    #fig.set_figwidth(len(numbers))
    #fig.set_figheight(len(values))
    fig.savefig(r"report/task_2.5_plot.pdf")

    plt.show()


def main(): # Функція для виклику усіх функцій :)

    path = Region()
    dataset = Calendar(path)
    Temperature(dataset)
    temperature_regime(dataset)
    Rose_winds(dataset)
    wind_activity(dataset)


if __name__ == "__main__": # Умова для виклику функцій у потрібній послідосності

    main() # Викликаємо усі функції