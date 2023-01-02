# -*- coding: utf-8 -*-


import openpyxl
from rich import print
import matplotlib.pyplot as plt
import eel



def sum_lists(list1, list2):

    return_list = []

    for i in range(len(list1)):

        return_list.append(list1[i] + list2[i])

    return return_list


def get_baza(baza_file):

    book = openpyxl.open(baza_file)
    sheet = book.active

    baza = []

    for row in range(2, sheet.max_row + 1):

        baza.append(sheet[row][0].value)

    return baza


def search_index_elements(element, list):

    index = [i for i, ltr in enumerate(list) if ltr == element]

    return index


def int_list(list):

    return_list = []

    for i in list:

        try:
            
            return_list.append(int(i))

        except:

            return_list.append(float(i))

    return return_list


def baza_info(baza_file):

    book = openpyxl.open(baza_file)
    sheet = book.active

    baza = []

    for row in range(2, sheet.max_row + 1):

        baza.append(sheet[row][1].value)

    return baza


@eel.expose
def get_user_info(data):

    book = openpyxl.open(data)
    sheet = book.active

    devices = []
    hours = []
    days = []

    for row in range(2, sheet.max_row + 1):

        devices.append(sheet[row][1].value)
        hour = sheet[row][2].value
        hour = hour.replace(",", ".")
        hours.append(hour)
        days.append(sheet[row][0].value)

    user_info = "День/Пристрої/Відповідний час використання(у годинах) <br>"

    for i in range(len(devices)):

        user_info += f"{i+1}. {days[i]}: {devices[i]} > {hours[i]} <br>"

    return user_info


@eel.expose
def replace_info(data, day, replace_text_device, replace_text_hours):

    book = openpyxl.open(data)
    sheet = book.active

    day += 1

    sheet[day][1].value = replace_text_device
    sheet[day][2].value = replace_text_hours

    book.save(data) # Зберігаємо файл
    book.close()


def devices_and_time(data):

    baza = get_baza(r"data/Devices.xlsx")

    book = openpyxl.open(data)
    sheet = book.active

    devices = []
    hours = []

    for row in range(2, sheet.max_row + 1):

        device = sheet[row][1].value
        #devices.append(sheet[row][1].value)
        hour = sheet[row][2].value
        hour = hour.replace(",", ".")
        #hours.append(hour)
        device = device.split()
        hour = hour.split()

        for i in range(len(device)):

            devices.append(device[i])
            hours.append(hour[i])

    hours = int_list(hours)
    hours_devices = []
    #print(devices, hours)
    for i in baza:

        index = search_index_elements(i, devices)
        time_element = 0

        for j in index:

            time_element += hours[j]

        hours_devices.append(time_element)

    #print(baza, hours_devices)

    return hours_devices


@eel.expose
def graph_users():

    users = ["Bogdan", "Maxim", "Oleg", "Pavel", "Vlad"]
    joint_list = [0, 0, 0, 0, 0, 0, 0]

    for user in users:

        data = f"data/Users/{user}.xlsx"
        book = openpyxl.open(data)
        sheet = book.active

        hours = []

        for row in range(2, sheet.max_row + 1):

            hour = sheet[row][2].value
            hour = hour.replace(",", ".")

            hour = hour.split()
            time = 0
            hour = int_list(hour)

            for i in range(len(hour)):

                time += hour[i]

            hours.append(time)

        joint_list = sum_lists(joint_list, hours)
        y = hours
        x = list(range(1, len(y) + 1))
        #print(x, y)
        fig, ax = plt.subplots()

        ax.bar(x, y) # створюємо стовпцеву діаграму
        plt.title("Діаграмa електричного наватаження")
        plt.xlabel("День (1 - Понеділок, 7 - неділя)")
        plt.ylabel("Час використання електроенергії (год)")
        ax.set_facecolor("seashell")
        fig.set_facecolor("floralwhite")
        fig.savefig(f"report/{user}_plot.pdf")
        fig.savefig(f"report/{user}_plot.png")
        fig.savefig(f"web/images/{user}_plot.png")

        #plt.show()

    y = joint_list
    x = list(range(1, len(joint_list) + 1))
    #print(x, y)
    
    fig, ax = plt.subplots()
    ax.bar(x, y) # створюємо стовпцеву діаграму
    plt.title("Спільна діаграмa електричного наватаження")
    plt.xlabel("День (1 - Понеділок, 7 - неділя)")
    plt.ylabel("Час використання електроенергії (год)")
    ax.set_facecolor("seashell")
    fig.set_facecolor("floralwhite")
    fig.savefig(f"report/all_plot.pdf")
    fig.savefig(f"report/all_plot.png")
    fig.savefig(f"web/images/all_plot.png")

    for i in range(len(joint_list)):

        joint_list[i] = float('{:.2f}'.format(joint_list[i]))

    #plt.show()
    return joint_list


@eel.expose
def main():

    data = r"data/Users/Bogdan.xlsx"
    users = ["Bogdan", "Maxim", "Oleg", "Pavel", "Vlad"]
    baza = get_baza(r"data/Devices.xlsx")
    days = ["Понеділок", "Вівторок", "Середа", "Чеьвер", "П'ятниця", "Субота", "Неділя"]
    d_a_t = [0, 0, 0, 0, 0, 0, 0]
    elect_info = baza_info(r"data/Devices.xlsx")
    info = graph_users()
    hour_days = ""
    d_a_t_info = ""
    ei = 0
    Tmax = 0
    odno = 90
    dvo = (90 + 45) / 2
    tre = (135 + 90 + 36) / 3

    for i in range(len(elect_info)):

        if elect_info[i] > ei:

            max_elem = baza[i]
            index_max_elem = i

    for user in users:
        
        data1 = f"data/Users/{user}.xlsx"
        d_a_t_i = devices_and_time(data1)

        if d_a_t_i[index_max_elem] > Tmax:

            Tmax = d_a_t_i[index_max_elem]

        d_a_t = sum_lists(d_a_t, d_a_t_i)

    for i in range(len(baza)):

        d_a_t_info += f"<br>{baza[i]}: {d_a_t[i]} кВт;"

    for i in range(len(info)):

        hour_days += f"<br>{days[i]}: {info[i]} кВт;"

    result1 = f"Загальне використання енергії пристроями: {d_a_t_info}<br>"
    result2 = f"Oбсяги споживання окремо для кожного дня тижня: {hour_days}<br>"
    result3 = f"Cумарні обсяги споживання за тиждень: {sum(info)} кВт;<br>"
    result4 = f"Пікове значення споживання електрики на день: {max(info)} кВт;<br>"
    result5 = f"Середнє значення споживання електрики на день: {(sum(info) / len(info)) :.2f} кВт;<br>"
    result6 = f"Tривалість використання максимального навантаження: {Tmax} годин;<br>"
    result7 = f"Cтупінь нерівномірності ГЕН: {((sum(info) / len(info)) / max(info)) :.2f};<br>"
    result8 = f"коефіцієнт використання встановленої потужності: {((sum(info) / len(info)) / elect_info[index_max_elem]) / 100 :.2f} годин;<br>"
    result9 = f"фінансові витрати за умов використання однозонного тарифу на електричну енергію: {sum(d_a_t) * (odno / 7) :.2f} Грн;<br>"
    result10 = f"фінансові витрати за умов використання двозонного тарифу на електричну енергію: {sum(d_a_t) * (dvo / 7) :.2f} Грн;<br>"
    result11 = f"фінансові витрати за умов використання трьозонного тарифу на електричну енергію: {sum(d_a_t) * (tre / 7) :.2f} Грн;<br>"
    result12 = f"фінансові витрати за умов використання однозонного тарифу на електричну енергію з оптимізацією ГЕН: {sum(d_a_t) * (odno * ((sum(info) / len(info)) / max(info))/ 7) :.2f} Грн;<br>"
    result13 = f"фінансові витрати за умов використання двозонного тарифу на електричну енергію з оптимізацією ГЕН: {sum(d_a_t) * (dvo * ((sum(info) / len(info)) / max(info))/ 7) :.2f} Грн;<br>"
    result14 = f"фінансові витрати за умов використання трьозонного тарифу на електричну енергію з оптимізацією ГЕН: {sum(d_a_t) * (tre * ((sum(info) / len(info)) / max(info))/ 7) :.2f} Грн;<br>"

    result = result1 + result2 + result3 + result4 + result5 + result6 + result7 + result8 + result9 + result10 + result11 + result12 + result13 + result14

    return result


if __name__ == "__main__":

    print(main())

    eel.init("web")

    eel.start("main.html", size = (1000, 1000))