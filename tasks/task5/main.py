# -*- coding: utf-8 -*-


import openpyxl
from rich import print
import matplotlib.pyplot as plt
import random
# Імпортуємо необхідні модулі


def delete(data):

    book = openpyxl.open(data) # Відкриваємо Excel файл тільки для читання
    sheet = book.active # вибираємо сторінку
  
    while(sheet.max_row > 1): # Проходимось по рядкам

        sheet.delete_rows(2) # Видаляємо усі дані

    book.save(data) # Зберігаємо файл
    book.close() # Завершуємо роботу з файлом та закриваємо його


def fill_random_info(data, lenght):

    delete(data) # очищуємо старі дані у файлі перед записом нових

    book = openpyxl.open(data) # Відкриваємо Excel файл тільки для читання
    sheet = book.active # вибираємо сторінку

    for row in range(2, lenght): # Заповнюємо дані від 2 рядка до введеного

        speed = random.randint(0, 25) # Генеруємо рандомне число для швидкості вітру у діапазоні від 0 до 25
        height = 10 # Задаємо одинакову сисоту башти
        coefficient = random.uniform(0, 0.5) # Генеруємо рандомне число для коефіцієнту потужності ВЕУ
        coefficient = float('{:.2f}'.format(coefficient)) # Обрізаємо число на 2 значення після крапки
        power = speed * height * coefficient # Вираховуємо швидкість враховуючи допоміжні дані
        power = float('{:.2f}'.format(power)) # Обрізаємо число на 2 значення після крапки
        time = random.uniform(1, 50) # Задаєсо час у вибраному діапазоні із типом данних float
        time = float('{:.2f}'.format(time)) # Обрізаємо число на 2 значення після крапки

        sheet[row][0].value = speed # Записуємо всі обраховані данні у файл 
        sheet[row][1].value = time # із кожною ітерацією записуємо нові дані у новий рядок
        sheet[row][2].value = power
        sheet[row][3].value = coefficient
        sheet[row][4].value = height

    book.save(data) # Зберігаємо файл
    book.close() # Завершуємо роботу з файлом та закриваємо його


def graph_info(data, save_name):

    book = openpyxl.open(data) # Відкриваємо Excel файл тільки для читання
    sheet = book.active # вибираємо сторінку

    power = []
    speed = []
    coefficient = []

    for row in range(2, sheet.max_row + 1):

        power.append(sheet[row][2].value) 
        speed.append(sheet[row][0].value)
        coefficient.append(sheet[row][3].value) # Получаємо потрібні значення із файлу

    speed.sort()
    power.sort()
    coefficient.sort() # Сортуємо значення
    #coefficient.reverse()

    #print(speed, power)
    fig, ax = plt.subplots() # Створюємо фігуру
    ax1 = ax.twinx() # Додаємо нову віст У
    ax.plot(speed, power, color = "r")#, marker = "o", label = "Енергія/швидкість")
    ax1.plot(speed, coefficient, color = "b")#, marker = ">", label = "Коеф.Потужн/швидкість")
    ax.set_ylabel("Виробляйма енергія (kW)") # Підписуємо вісь У зліва
    ax1.set_ylabel("коефіцієнту потужності ВЕУ") #  Підписуємо вісь У зправа
    plt.xlabel("Швидкість вітру (м/c)") # Підписуємо спільну вісь У
    fig.savefig(save_name) # Зберігаємо графік у пдф за шляхом
    plt.show() # Демонструємо графік


def correct_data(data):

    book = openpyxl.open(data) # Відкриваємо Excel файл тільки для читання
    sheet = book.active # вибираємо сторінку

    #print("[bold green]Ведіть висоту башти ВЕУ в м (висота дана у файлі - 10м)[/bold green]")
    height_s = 12 #int(input("Введіть висоту: ")) # Задаємо висоту 


    for row in range(2, sheet.max_row + 1):

        power = sheet[row][2].value
        height_w = sheet[row][4].value

        new_power = power * (height_s / height_w) ** 0.14 # Обраховуємо новий виробіток енергії на зміненій висоті за формулою
        new_power = float('{:.2f}'.format(new_power)) # Обрізаємо до 2 заначень після крапки

        sheet[row][4].value = height_s # Записуємо нові значення у файл
        sheet[row][2].value = new_power

    book.save(data) # Зберігаємо файл
    book.close() # Завершуємо роботу з файлом та закриваємо його


def all_energy(data):

    book = openpyxl.open(data) # Відкриваємо Excel файл тільки для читання
    sheet = book.active # вибираємо сторінку

    time = []
    power = []

    for row in range(2, sheet.max_row + 1):

        power.append(sheet[row][2].value)
        time.append(sheet[row][1].value)

    time_and_power = []

    for i in range(len(time)):

        t_a_p = time[i] * power[i] # Обраховуємо кількість виробленої енергії за весь час
        time_and_power.append(t_a_p)

    energy = sum(time_and_power) # Отримуємо повний виробіток врахувавши час
    energy = float('{:.2f}'.format(energy)) # обрізаємо на 2 значення після крапки

    return energy


def main():

    data = r"data/data.xlsx" # Шлях до бази данних
    price = 7 # Ціна за 1 кВт по зеленому тарифу
    CO2 = all_energy(data) * 10**-3 * 1.06 # Розраховуэмо оцінку обсягів скорочень викидів парникових газів у тонах СО2 еквіваленту
    lenght = 30 # Кількість рядів данних у файлі

    fill_random_info(data, lenght) # Заповнюємо базу рандомними даними
    print("[bold blue]Синія лінія на графіку - відношення швидості і виробленої енергії (ось У ліворуч) [/bold blue]")
    print("[bold red]Червона лінія на графіку - відношення швидості і поткужності ВЕУ (ось У праворуч) [/bold red]")
    graph_info(data, r"report/start_data_plot.pdf") # Демонструємо дані як у методичці
    correct_data(data) # Корегуємо дані вводячи висоту Башти 
    graph_info(data, r"report/and_data_plot.pdf") # Демонструємо дані із відкорегованою висотою
    print(f"[yellow]Обяги згенерованої енергії за визначений проміжок часу: {all_energy(data)} кВт;[/yellow]")
    print(f"[yellow]Oцінкa обсягів скорочень викидів парникових газів у тонах СО2 еквіваленту: {CO2 :.2f} т;[/yellow]")
    print(f"[yellow]Дохід від продажу електричної енергії за «зеленим» тарифом: {all_energy(data) * price * 10 :.2f} грн;[/yellow]")
    print(f"[yellow]Дохід від продажу одиниць скорочення викидів (ОСВ): {CO2 * 10 * 15 :.2f} € ({CO2 * 10 * 15 * 35:.2f} грн);[/yellow]")
    print("[bold green]Графіки збережено у папку 'reports' [/bold green]")


if __name__ == "__main__":

    main()